import pathlib
import re
from typing import Literal
from pandas._typing import Scalar
import openpyxl
import pandas as pd
import tkinter as tk
import os
from tkinter import filedialog
import pdfplumber
from docx import Document
from docx.shared import Pt
from docxtpl import DocxTemplate
from openpyxl.utils.dataframe import dataframe_to_rows



with open(r'codes_dict.py', 'r', encoding='utf-8') as file:
    file_content = file.read()

# Выполнение содержимого файла
namespace = {}
exec(file_content, namespace)

# Теперь в переменной namespace будет доступен ваш словарь
bar_codes: dict[str, str] = namespace.get('bar_codes', {})


def extract_data_from_tableOtherType(table: list[list[str | None]]) -> list[tuple[str, str, str]]:
    data: list[tuple[str, str, str]] = []

    for row in table:
        if isinstance(row, list):
            # Преобразуем список в строку для поиска по регулярному выражению
            row_str = ', '.join(map(str, row))
            matches = re.findall(r'([^,]+).*?((?:CLNW|CSNW)\d+).*?(\d+)', row_str)

            if matches:
                # name = matches[0][0].strip()
                code = matches[0][1]
                try:
                    name = bar_codes[code]
                except:
                    name = ''
                if name:
                    quantity = matches[0][2]
                    quantity_parts = quantity.split(',')
                    quantity_integer = int(quantity_parts[0])
                    for _ in range(quantity_integer):
                        data.append((code, name, f"{quantity_integer},00"))


    return data

def extract_data_from_table(table: list[list[str | None]], system_transfer: bool = False, true_data: bool = False) -> list[tuple[str, str, str]]:
    data: list[tuple[str, str, str]] = []
    next_row_has_quantity = False  # Флаг, указывающий, что следующая строка содержит количество
    quantity_position: list[str|None] = []
    for idx, row in enumerate(table):
        try:
            if isinstance(row, list):
                row_str = ', '.join(map(str, row))
                matches = re.findall(r'((?:CLNW|CSNW)\d+).*?(?:, ([^,]+)|(?= \d{6}\b))', row_str)
                quantity = ""
                if matches:
                    matches
                # проверяем первую строку на наличие количества
                # first_row = table[idx + 1]
                # first_row_str = ', '.join(map(str, first_row))
                # quantity_match = re.search(r', (\d+,\d+)', first_row_str)
                # if not quantity_position:
                #     quantity_position = [i for i in range(len(table[idx])) if table[idx][i].replace("\n", '').lower() == 'кол-возапрошено' or table[idx][i].replace("\n", '').lower() == 'кол-во запрошено']
                # if quantity_position:
                #     quantity = first_row[quantity_position[0]].replace("\n", '')
                #     next_row_has_quantity = True
                # if not next_row_has_quantity:
                if not system_transfer and ((matches and len(matches[0][1]) > 3) or ("Поставщик" in row_str or "запроше" in row_str)):
                    # Проверяем следующую строку на наличие количества
                    if idx + 1 < len(table):
                        next_row = table[idx + 1] #idx + 1
                        # if not quantity_position:
                        #     quantity_position = [i for i in range(len(table[idx])) if table[idx][i].replace("\n", '').lower() == 'кол-возапрошено' or table[idx][i].replace("\n", '').lower() == 'кол-во запрошено']
                        next_row_str = ', '.join(map(str, next_row))
                        quantity_match = re.search(r', (\d+,\d+)', next_row_str)
                        if quantity_match:
                            quantity = quantity_match.group(1)
                            # quantity = next_row[quantity_position[0]].replace("\n", '')
                            next_row_has_quantity = True
                        else:
                            next_row_has_quantity = False
                    if not next_row_has_quantity:
                        curr_row_str = ', '.join(map(str, table[idx]))
                        quantity_match = re.search(r', (\d+,\d+)', curr_row_str)
                        if quantity_match:
                            next_row_has_quantity = True
                else:
                    first_row = table[idx]
                    # first_row_str = ', '.join(map(str, first_row))
                    # quantity_match = re.search(r', (\d+,\d+)', first_row_str)
                    if not quantity_position:
                        quantity_position = [i for i in range(len(table[idx])) if table[idx][i].replace("\n", '').lower() == 'кол-возапрошено' or table[idx][i].replace("\n", '').lower() == 'кол-во запрошено']
                    if quantity_position:
                        quantity = first_row[quantity_position[0]].replace("\n", '')
                        next_row_has_quantity = True
                if matches and not quantity and not true_data and not system_transfer:
                    if matches:
                        if matches[0][0] in bar_codes:
                            code = matches[0][0]
                            name = bar_codes[code]
                            return [data, code, name]
                elif quantity and not matches and not true_data and not system_transfer:
                    return [data, quantity]
                
                if not quantity:
                    quantity = "1,00"
                
                for match in matches:
                    code = match[0]
                    # name = match[1]
                    name = bar_codes[code]
                    if code and name and code in bar_codes.keys():
                        if next_row_has_quantity:
                            # if 'кол' in quantity.lower():
                            #     if idx+1 < len(table):
                            #         quantity = table[idx+1][2]
                            quantity_parts = quantity.split(',')
                            quantity_integer = int(quantity_parts[0])
                            for _ in range(quantity_integer):
                                data.append((code, name, f"{quantity_integer},00"))
                        else:
                            data.append((code, name, "1,00"))
        except Exception as e:
            print(f"ex {str(e)}")
    return data

    
def extract_data_from_excelTable(excel_filePath: str) -> tuple[list, Scalar, Scalar]:
    # Заголовки
    # header = ['№п/п', 'Номер актива', 'Наименование актива', 'Серийный номер Запасов', 'Код позиции Запасов', 'Ед. измерения', 'Кол-во']

    # # Сразу задаем заголовки при создании DataFrame
    # try:
    #     if pd.read_excel(excel_filePath).iloc[:,8].notnull().any():
    #         header = ['№п/п', 'Номер актива', 'Наименование актива', 'Серийный номер Запасов', 'Код позиции Запасов', 'Ед. измерения', 'Кол-во', '']
    # except:
    #     pass
    df = pd.read_excel(excel_filePath) #names=header
    # Извлекаем адреса БС из строки ? и столбцов B и E
    bsFromName = df.iloc[8, 1] 
    bsToName = df.iloc[8, 4]  

    # Инициализируем начальные значения
    table_started = False
    table_data = []

    for index, row in df.iterrows():
        row_str = ', '.join(map(str, row))
        matches = re.findall(r'((?:CSCM|CSIT|CSNW)\d+).*?(?:, ([^,]+)|(?= \d{6}\b))', row_str)

        if matches:
            if not table_started:
                # Начало таблицы
                table_started = True
                table_data = []
            table_data.append(row)

    return table_data, bsFromName, bsToName


def create_data_table(data: list[list], countTable1: int) -> tuple[list[dict[str, str]], int]:
    table_data: list[dict[str, str]] = []
    for idx, item in enumerate(data):
        try:
            countTable1 += 1
            row = {
                "N": countTable1,  # Номер строки
                "P": item[1],  # Наименование оборудования
                "D": "1",  # Количество , поставил 1 что бы всегда выходило
                "M": "шт",  # Единицы измерения
                "C": "",  # Серийный номер (пусто)
                "S": item[0],  # Код запаса ESS/НФС
                "T": "",  # Примечания (пусто)
            }
            table_data.append(row)
        except:
            print('Ошибка на строке 170: ', item[0])
    return table_data , countTable1

def transform_list(input_list: list)  -> list[list]:
    transformed_list = [[item[5], item[3]] for item in input_list] #item[5] Код, item[3] Найменование
    return transformed_list

def insert_table_into_template(tables: list, pdf_filePath: str, bs_code: str|None, second_type_table: list, bs_name:str|None, third_type_table: list, pdf_fileName: str|None, bs_name_otv: str|None, bs_code_otv: str|None) -> None:
    template_path = "АТП_актШаблон.docx"

    doc = DocxTemplate(template_path)
    if not bs_name:
        bs_name = bs_name_otv
    context = {"DATA_FIRST": tables, "BS_NAME": bs_name, "BS_CODE": bs_code, "SECOND_TYPE_TABLE": second_type_table, "THIRD_TYPE_TABLE": third_type_table, "BS_CODE_OTV": bs_code_otv, "BS_NAME_OTV": bs_name_otv}
    
    doc.render(context)
    
    word_filePath = os.path.join(os.path.dirname(os.curdir), f"Акт_монтажа_{bs_name}.docx")
    
    doc.save(word_filePath)
    print(f"Готово: {word_filePath}")

def clean_text(text):
    # Разделяем текст по \n
    parts = text.split('\n')
    
    # Если есть две части, берем текст от первой до второй
    # if len(parts) >= 2:
    #     cleaned_text = parts[1]
    # else:
    # Если нет второй части, берем всю строку
    cleaned_text = parts[0]
    cleaned_code = parts[1]
    # Разделяем текст по запятой
    parts = cleaned_code.split(',')
    
    # Если есть вторая часть, берем ее (это может быть код)
    if len(parts) >= 2:
        # cleaned_text = parts[0].strip()  # Берем текст до запятой
        code = parts[0].strip()
    else:
        code = None
    
    return cleaned_text, code


def getDataFromFirstTable(table):
    bs_cleaned_text,bs_code,num_doc,date_doc = None ,None,None,None
    
    # Преобразуем таблицу в DataFrame
    df = pd.DataFrame(table[2:], columns=table[1])

    # Получаем данные из 4-го столбца и добавляем в список
    bs_data = df.iloc[:, 3].tolist()
    num_doc_data = df.iloc[:, 0].tolist()
    date_doc_data = df.iloc[:, 1].tolist()

    # Обрабатываем каждую строку
    for i in range(len(bs_data)):
        bs_cleaned_text, bs_code = clean_text(bs_data[i])
        num_doc = num_doc_data[i]
        date_doc = date_doc_data[i]

        print(bs_cleaned_text, bs_code, num_doc, date_doc)
        # data.append({'cleaned_text': bs_cleaned_text, 'code': bs_code, 'num_doc': num_doc, 'date_doc': date_doc})
    return bs_cleaned_text,bs_code,num_doc,date_doc

def main_handle_of_pdf_files(files_folder: tuple[str, ...] | Literal[''], second_type_table: list[dict[str, str]], third_type_table: list[dict[str, str]], tables: list[dict[str, str]], is_otv_file: bool = False)-> tuple[list[dict[str, str]], str, str, list[dict[str, str]], str , list[dict[str, str]], str]:
    count: int = 0 
    countTable = 0
    bs_code = ''
    bs_name = ''
    bsToName = ""
    for file_path in files_folder:
        pdf_filePath: str = file_path
        file_name = re.search('([^\\/]+)$', file_path)[1]
        if file_name.endswith('.pdf'):
            data = []
            invoice_number = ""
            date = ""
            bs_flag = True
            system_transfer = False
            if 'перенос' in file_name.lower():
                system_transfer = True

            with pdfplumber.open(pdf_filePath) as pdf:
                old_format = False
                if 'Накладная №' in pdf.pages[0].extract_text():
                    old_format = True
                temp_list = []
                for page in pdf.pages:
                    pdf_tables = page.extract_tables()
                    for table in pdf_tables:
                        dataMid = []
                        temp_data = extract_data_from_table(table, system_transfer)
                        if len(temp_data) > 1 and isinstance(temp_data[1], str):
                            if len(temp_data) == 3:
                                temp_list.extend(temp_data[1:])
                            if len(temp_data) == 2:
                                temp_list.append(temp_data[1])
                            if len(temp_list) == 3:
                                # if temp_data[0]:
                                    # data.extend(temp_data[0])
                                temp_data = []
                                for i in range(int(temp_list[2].split(',')[0])-1):
                                    temp_data.append((temp_list[0], temp_list[1], '1,00'))
                                data.extend(temp_data)
                                temp_list = []
                                
                        dataMid.extend(extract_data_from_table(table, system_transfer))
                        if dataMid:
                            data.extend(extract_data_from_table(table, system_transfer, True)) 
                        else: 
                            data.extend(extract_data_from_tableOtherType(table))
                # Извлекаем bs_name, bs_code, номер накладной и дату из первой страницы
                first_page = pdf.pages[0]
                text = first_page.extract_text()
                print(file_name)
                if old_format:
                    bs_name_match = re.search(r'СП Получатель:\s*.*?\.(\d+).(.*?)-\d+', text)
                    invoice_number_match = re.search(r'Накладная № (\d+)', text)
                    cleaned_text = re.sub(r'[^\d.]', '', text)

                    date_match = re.search(r'(\d{2}\.\d{2}.\d{4})', cleaned_text)

                    bs_code_match = re.search(r'СП Получатель:.*?-(\d+)-', text)
                    if bs_name_match:
                        bs_name = bs_name_match.group(2)
                        bs_code = bs_name_match.group(1)
                    else:
                        bs_name = ''
                        bs_code = ''
                        
                    if invoice_number_match:
                        invoice_number = invoice_number_match.group(1)
                    if date_match:
                        date = date_match.group(1)
                    if bs_code_match:
                        bs_code = bs_code_match.group(1)
                else:
                    pdf_table = pdf.pages[0].extract_tables()[0]
                    bs_name, bs_code, invoice_number, date  = getDataFromFirstTable(pdf_table)
                    # bs_flag = False

                
                    
                count += 1
                if bs_flag:
                    textForO = f"Оборудование получено по накладной на БС {bs_name}"
                else:
                    textForO = f"Оборудование отгружено по накладной на внутреннее перемещение {bs_name}"
                    # invoice_number = ''
                    # date = ''

                row = {
                    "N": count,  # Номер строки
                    "O": textForO,
                    "I": invoice_number,
                    "D": date,
                    "T": ""
                }

                second_type_table.append(row)
                
                table_data, countTable = create_data_table(data, countTable)

                # Удаляем символы переноса строки и заменяем их на пробелы в тексте второго столбца
                for item in table_data:
                    if item["P"]:
                        item["P"] = re.sub(r'\n|\r', ' ', item["P"])

                # Добавляем данные в общий список
                if is_otv_file:
                    third_type_table.extend(table_data)
                else:
                    tables.extend(table_data) 
        elif file_path.endswith('.xlsx') and file_path.__contains__("Накладная" or "накладная"):
            
            excel_filePath = file_path
            data,bsFromName,bsToName = extract_data_from_excelTable(excel_filePath)

            if str(bsToName) == "nan":
                print(f"Другой шаблон файла: {file_name}")   
                

            count += 1
            row = {
                "N": count,  # Номер строки
                "O": f"Накладная на перемещение c {bsFromName} на БС {bsToName}",
                "I": "",
                "D": "",
                "T": ""
            }
            
            second_type_table.append(row)

            dataForTable: list[list] = transform_list(data)

            table_data,countTable = create_data_table(dataForTable, countTable)
            tables.extend(table_data)
            
    if not bs_name and bsToName:
        bs_name = bsToName
    return tables, pdf_filePath, bs_code, second_type_table, bs_name, third_type_table, file_name

def startProcessPdfToWord(selected_files: list[str], selected_files_otv: list[str]|None) -> None:
    tables = []  
    second_type_table = []
    third_type_table = []

    
    bs_name = ""
    bs_code = ""
    bs_name_otv = ""
    bs_code_otv = ""
    
    tables, pdf_filePath, bs_code, second_type_table, bs_name, third_type_table, pdf_fileName = main_handle_of_pdf_files(selected_files, second_type_table, third_type_table, tables)
    if selected_files_otv:
        tables, pdf_filePath, bs_code_otv, second_type_table, bs_name_otv, third_type_table, pdf_fileName = main_handle_of_pdf_files(selected_files_otv, second_type_table, third_type_table, tables, is_otv_file=True)
    insert_table_into_template(tables, pdf_filePath, bs_code, second_type_table, bs_name, third_type_table, pdf_fileName, bs_name_otv, bs_code_otv)
    


class FileSelector:
    def __init__(self) -> None:
        self.selected_files_otv: list[tuple[str, ...] | Literal['']] = []
        self.selected_files: list[tuple[str, ...] | Literal['']] = []

def selecting_some_files() -> None:
    FileSelector.selected_files = filedialog.askopenfilenames(title = "Выберите файлы") #, filetypes = (("PDF files", "*.pdf"), ("all files", "*.*"))
    FileSelector.selected_files_otv = []
    # return files

def selecting_some_otv_files() -> None:
    FileSelector.selected_files_otv = filedialog.askopenfilenames(title = "Выберите ответные файлы") #, filetypes = (("PDF files", "*.pdf"), ("all files", "*.*"))
    # return files
def start_handle() -> None:
    startProcessPdfToWord(FileSelector.selected_files, FileSelector.selected_files_otv)
    

def fill_dictionary() -> None:
    new_dicts = filedialog.askopenfilename(title = "Выберите Excel", filetypes = (("Excel files", "*.xlsx"), ("all files", "*.*")))
    new_dicts_workbook = openpyxl.load_workbook(new_dicts)
    new_dicts_worksheet = new_dicts_workbook.active
    for row in new_dicts_worksheet.iter_rows(min_row=2, max_row=new_dicts_worksheet.max_row, min_col=1, max_col=2): 
        bar_codes[row[0].value] = row[1].value
        print(f"Добавлено: '{row[0].value}': '{row[1].value}'")

    with open('codes_dict.py', 'w', encoding='utf-8') as file:
        file.write(f"bar_codes: dict[str, str] = {bar_codes}")
